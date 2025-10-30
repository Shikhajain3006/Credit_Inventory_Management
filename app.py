"""
AI-Powered Credit Memo SOX Automation Tool - Streamlit Version
Same as ai-app but for credit memo validation
Version: 2.0 - Fixed SOX Status logic and chat interface
"""

import streamlit as st
import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import re

from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import HumanMessage
import matplotlib.pyplot as plt

# ===================== Setup =====================

# Load environment variables - from parent directory
# Use absolute path directly to handle Streamlit's working directory quirks
app_dir = Path(__file__).resolve().parent
env_path = app_dir.parent / ".env"

# First try load_dotenv
load_dotenv(env_path, override=True)

# For Streamlit, also read directly from file to ensure variables are set
if env_path.exists():
    with open(env_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                # Force set in os.environ
                os.environ[key] = value

# Debug: Verify environment loaded
DEBUG_MODE = False  # Set to True to see debug info
if DEBUG_MODE:
    print(f"ENV PATH: {env_path}")
    print(f"ENV EXISTS: {env_path.exists()}")
    print(f"API KEY LOADED: {bool(os.getenv('AZURE_OPENAI_API_KEY'))}")
    print(f"API KEY VALUE (first 20 chars): {os.getenv('AZURE_OPENAI_API_KEY', '')[:20] if os.getenv('AZURE_OPENAI_API_KEY') else 'NOT SET'}")
    print(f"ENDPOINT LOADED: {bool(os.getenv('AZURE_OPENAI_ENDPOINT'))}")
    print(f"ENDPOINT LOADED: {bool(os.getenv('AZURE_OPENAI_ENDPOINT'))}")

# Page config
st.set_page_config(
    page_title="Credit Memo SOX Automation",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .header { color: #1f77b4; font-size: 2em; font-weight: bold; }
    .success { color: #2ca02c; }
    .warning { color: #ff7f0e; }
    .error { color: #d62728; }
    </style>
""", unsafe_allow_html=True)

# Initialize AI client (NOT cached - needs fresh check each time)
def get_ai_client():
    try:
        # Try to get from environment first, then from Streamlit secrets
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        
        # Try to get from secrets if not in env
        try:
            if not api_key:
                api_key = st.secrets.get("AZURE_OPENAI_API_KEY")
            if not endpoint:
                endpoint = st.secrets.get("AZURE_OPENAI_ENDPOINT")
        except:
            pass  # Secrets file doesn't exist, continue with env vars
        
        if not api_key or not endpoint:
            return None
            
        return AzureChatOpenAI(
            azure_deployment="gpt-4.1",
            api_version="2025-01-01-preview",
            temperature=0.2,
            azure_endpoint=endpoint,
            api_key=api_key,
        )
    except Exception as e:
        return None

# ===================== AI Helper Functions =====================

def get_ai_response(client, context_prompt, user_query):
    """Get response from AI client"""
    try:
        if not client:
            return None
        
        full_prompt = f"""{context_prompt}

User Query: {user_query}

Please provide a helpful, concise response."""
        
        response = client.invoke([HumanMessage(content=full_prompt)])
        return response.content
    except Exception as e:
        return f"❌ Error: {str(e)}"

def build_context_prompt(result_df):
    """Build context prompt from validation results"""
    compliant = (result_df["SOX Status"] == "SOX Compliant").sum()
    violations = (result_df["SOX Status"] == "SOX Violation").sum()
    high_risk = (result_df["Risk Level"] == "High").sum()
    med_risk = (result_df["Risk Level"] == "Medium").sum()
    over_sla = (result_df["Timeline Status"].astype(str).str.startswith("Over")).sum()
    
    return f"""You are an expert SOX compliance analyst reviewing credit memo validation results.

Current validation results:
- Total Memos: {len(result_df)}
- Compliant: {compliant} ({100*compliant/len(result_df):.1f}%)
- Violations: {violations} ({100*violations/len(result_df):.1f}%)
- High Risk: {high_risk}
- Medium Risk: {med_risk}
- Over SLA: {over_sla}

Summary Data:
{result_df[['Memo', 'Customer Name', 'Amount', 'SOX Status', 'Risk Level', 'Reason Class', 'Timeline Status']].head(20).to_string()}"""

# ===================== Helper Functions =====================

def normalize_header_text(s):
    """Normalize text for comparison"""
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower())

def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names"""
    df = df.copy()
    df.columns = (
        df.columns.str.strip()
        .str.replace(r"\s+", " ", regex=True)
        .str.title()
    )
    return df

def find_header_row(raw: pd.DataFrame, must_have_tokens) -> Optional[int]:
    """Find header row automatically"""
    for i in range(min(30, len(raw))):
        vals = [normalize_header_text(v) for v in list(raw.iloc[i].values)]
        if all(any(tok in v for v in vals if isinstance(v, str)) for tok in must_have_tokens):
            return i
    return None

def parse_upper_from_amount_range(s: str) -> float:
    """Parse amount ranges"""
    if pd.isna(s):
        return float("inf")
    s_raw = str(s)
    s = s_raw.strip().replace(",", "")
    low = s.lower()
    if low.startswith("up to"):
        tail = low.split("up to", 1)[1]
        nums = re.sub(r"[^\d.]", "", tail)
        return float(nums) if nums else float("inf")
    if "–" in s or "-" in s:
        sep = "–" if "–" in s else "-"
        right = s.split(sep)[-1]
        nums = re.sub(r"[^\d.]", "", right)
        return float(nums) if nums else float("inf")
    if low.startswith("above"):
        return float("inf")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    return float(nums[-1]) if nums else float("inf")

def read_matrix_table(xl: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """Read and parse approval matrix"""
    raw = xl.parse(sheet_name, header=None)
    must_have = ["amount range", "approver level", "designation"]
    hdr = find_header_row(raw, must_have)
    if hdr is None:
        return None
    
    m = xl.parse(sheet_name, header=hdr)
    m = norm_cols(m).rename(columns={"Approver Level": "Level"})
    
    if not all(col in m.columns for col in ["Amount Range", "Level", "Designation"]):
        return None
    
    # Convert Level to numeric, handling various formats
    m["Level"] = pd.to_numeric(
        m["Level"].astype(str).str.extract(r"(\d+)")[0], 
        errors="coerce"
    ).astype("Int64")
    m["Upper Limit"] = m["Amount Range"].apply(parse_upper_from_amount_range)
    m = m.sort_values(["Upper Limit", "Level"]).reset_index(drop=True)
    return m[["Level", "Designation", "Upper Limit"]]

def classify_matrix_sheet_name(name: str) -> Optional[str]:
    """Classify matrix type by sheet name"""
    n = name.lower()
    if "promotional" in n:
        return "promotional"
    if "contract" in n:
        return "contract"
    if "other" in n:
        return "other"
    return None

def map_summary_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map column synonyms to canonical names"""
    col_map = {}
    norm_to_col = {normalize_header_text(c): c for c in df.columns}
    
    mappings = {
        "Memo": ["memo", "memo id", "memoid"],
        "Customer Name": ["customer name"],
        "Cm Date": ["cm date", "credit memo date"],
        "Created By": ["created by", "creator"],
        "Amount": ["amount"],
        "Reason": ["reason"],
        "Date Of Approval": ["date of approval", "approval date"],
        "Approver": ["approver"],
        "Approver Designation": ["approver designation", "designation"],
    }
    
    for canonical, syns in mappings.items():
        for syn in syns:
            if syn in norm_to_col:
                col_map[norm_to_col[syn]] = canonical
                break
    
    return df.rename(columns=col_map)

# ===================== Validation Engine =====================

class ValidateCredit:
    def __init__(self, sla_days=5, missing_levels_for_high=2, keywords_promotional=None, keywords_contract=None):
        self.sla_days = sla_days
        self.missing_levels_for_high = missing_levels_for_high
        self.keywords_promotional = keywords_promotional or ["promotional", "promotion"]
        self.keywords_contract = keywords_contract or ["contract"]
        self.ai_client = get_ai_client()
    
    def validate(self, df, matrices):
        """Validate all memos"""
        df["Reason Class"] = ""
        df["Required Approval Level"] = ""
        df["Final Approver"] = df["Approver Designation"]
        df["Final Approver Level"] = ""  # Will be filled with int or empty string, then converted to string
        df["SOX Status"] = ""
        df["Risk Level"] = ""
        df["Missing Approvals"] = ""
        df["Violation Reason"] = "None"  # Track violation reasons
        df["Violation Count"] = 0  # Track number of violations
        df["Approval Timeline (Business Days)"] = pd.NA
        df["Timeline Status"] = ""
        df["Approval Sequence"] = ""
        df["Designation Level Check"] = ""
        df["Duplicate Memo"] = ""
        
        # Check duplicates
        df["Duplicate Memo"] = df.duplicated(subset=["Memo"], keep=False).map({True: "Yes", False: "No"})
        
        # Validate each row
        for i, row in df.iterrows():
            self._validate_row(df, i, row, matrices)
        
        # Ensure Final Approver Level is string type (not mixed int/str)
        df["Final Approver Level"] = df["Final Approver Level"].astype(str)
        df.loc[df["Final Approver Level"] == "nan", "Final Approver Level"] = ""
        
        return df
    
    def _validate_row(self, df, i, row, matrices):
        """Validate single memo"""
        amt = pd.to_numeric(row["Amount"], errors="coerce")
        
        # Classify reason
        reason = str(row["Reason"]).lower()
        if any(k in reason for k in self.keywords_promotional):
            rc = "Promotional"
        elif any(k in reason for k in self.keywords_contract):
            rc = "Contract"
        else:
            rc = "Other"
        
        df.at[i, "Reason Class"] = rc
        mtx = matrices.get(rc.lower())
        
        # Required level
        need_lvl = None
        if mtx is not None and not pd.isna(amt):
            rows = mtx[mtx["Upper Limit"] >= amt].sort_values("Upper Limit")
            if not rows.empty:
                need_lvl = int(rows["Level"].iloc[0])
        
        df.at[i, "Required Approval Level"] = need_lvl if need_lvl else ""
        
        # Approver level - with fuzzy matching (3 strategies)
        apr_lvl = None
        desig = str(row["Approver Designation"]).strip()
        
        if mtx is not None and desig:
            dnorm = desig.casefold()
            
            # Strategy 1: Exact match
            m = mtx[mtx["Designation"].str.strip().str.casefold() == dnorm]
            if not m.empty:
                apr_lvl = int(m["Level"].iloc[0])
            # Strategy 2: Substring match (matrix designation contains user input)
            elif not mtx.empty:
                contains = mtx[mtx["Designation"].str.strip().str.casefold().str.contains(re.escape(dnorm))]
                if not contains.empty:
                    apr_lvl = int(contains["Level"].iloc[0])
                # Strategy 3: Reverse match (user input contains matrix designation)
                else:
                    reverse = mtx[mtx["Designation"].str.strip().str.casefold().apply(lambda x: x in dnorm)]
                    if not reverse.empty:
                        apr_lvl = int(reverse["Level"].iloc[0])
                    else:
                        apr_lvl = "NOT_FOUND"
        
        df.at[i, "Final Approver Level"] = apr_lvl if isinstance(apr_lvl, int) else ""
        
        # Check compliance - Track approval level status but defer final SOX Status
        approval_compliant = False
        
        if pd.isna(amt) or need_lvl is None:
            df.at[i, "SOX Status"] = "SOX Violation"
            df.at[i, "Risk Level"] = "High"
            df.at[i, "Missing Approvals"] = "Missing amount or matrix not available"
        elif apr_lvl is None:
            df.at[i, "SOX Status"] = "SOX Violation"
            df.at[i, "Risk Level"] = "High"
            df.at[i, "Missing Approvals"] = "Approver designation missing"
        elif apr_lvl == "NOT_FOUND":
            df.at[i, "SOX Status"] = "SOX Violation"
            df.at[i, "Risk Level"] = "High"
            df.at[i, "Missing Approvals"] = f"Designation '{row['Approver Designation']}' not found in matrix"
        elif apr_lvl >= need_lvl:
            # Approval level is sufficient, but don't set SOX Compliant yet
            # Timeline must also be checked before final determination
            approval_compliant = True
            df.at[i, "SOX Status"] = ""  # Placeholder - will be set after timeline check
            df.at[i, "Risk Level"] = ""
            df.at[i, "Missing Approvals"] = "None"
        else:
            missing = list(range(apr_lvl + 1, need_lvl + 1))
            risk = "High" if len(missing) >= self.missing_levels_for_high else "Medium"
            df.at[i, "SOX Status"] = "SOX Violation"
            df.at[i, "Risk Level"] = risk
            df.at[i, "Missing Approvals"] = f"Level {'–'.join(map(str, missing))} Missing"
        
        # Timeline: Approval -> CM within SLA
        # Logic: Approval comes first, then CM should be created within SLA
        cm = pd.to_datetime(row["Cm Date"], errors="coerce")
        ap = pd.to_datetime(row["Date Of Approval"], errors="coerce")
        
        if pd.isna(cm) or pd.isna(ap):
            df.at[i, "Approval Timeline (Business Days)"] = pd.NA
            df.at[i, "Timeline Status"] = "Dates Missing"
            df.at[i, "Approval Sequence"] = "Dates Missing"
            # If approval was compliant but dates are missing, mark as violation
            if approval_compliant:
                df.at[i, "SOX Status"] = "SOX Violation"
                df.at[i, "Risk Level"] = "High"
                df.at[i, "Missing Approvals"] = "Timeline: Dates missing"
        else:
            # Check if approval came before CM (correct order)
            if ap > cm:
                # Approval happened AFTER CM was created - this is a violation
                bdays = pd.bdate_range(cm, ap).size - 1
                df.at[i, "Approval Timeline (Business Days)"] = bdays
                df.at[i, "Timeline Status"] = "Approval After CM"
                df.at[i, "Approval Sequence"] = "Approval After CM (Violation)"
                # Sequence violation takes priority
                df.at[i, "SOX Status"] = "SOX Violation"
                df.at[i, "Risk Level"] = "High"
                df.at[i, "Missing Approvals"] = "Approval Date: Approved after CM creation"
            else:
                # Approval came before CM (correct order) - check SLA
                bdays = pd.bdate_range(ap, cm).size - 1
                df.at[i, "Approval Timeline (Business Days)"] = bdays
                
                if bdays <= self.sla_days:
                    df.at[i, "Timeline Status"] = f"Within {self.sla_days} days"
                    df.at[i, "Approval Sequence"] = "Order OK"
                    # Only mark as compliant if approval was also compliant AND timeline is OK
                    if approval_compliant:
                        df.at[i, "SOX Status"] = "SOX Compliant"
                        df.at[i, "Risk Level"] = "Low"
                else:
                    # SLA violated
                    df.at[i, "Timeline Status"] = f"Over {self.sla_days} days"
                    df.at[i, "Approval Sequence"] = "SLA Violated"
                    # SLA violation overrides approval compliance
                    df.at[i, "SOX Status"] = "SOX Violation"
                    df.at[i, "Risk Level"] = "Medium"
                    df.at[i, "Missing Approvals"] = f"Timeline: CM created {bdays - self.sla_days} days after SLA threshold"
        
        # Generate Violation Reason summary
        violation_reasons = []
        
        # Check for approval level violations
        missing_approvals = df.at[i, "Missing Approvals"]
        if df.at[i, "SOX Status"] == "SOX Violation" and missing_approvals and missing_approvals != "None":
            if "Level" in str(missing_approvals):
                violation_reasons.append(f"Missing Approval: {missing_approvals}")
            elif "Timeline" in str(missing_approvals):
                violation_reasons.append(f"SLA Breach: {missing_approvals}")
            else:
                violation_reasons.append(f"Approval Issue: {missing_approvals}")
        
        # Check for timeline violations (from Approval Sequence)
        approval_seq = df.at[i, "Approval Sequence"]
        if "SLA Violated" in str(approval_seq):
            violation_reasons.append(f"SLA Exceeded: {df.at[i, 'Timeline Status']}")
        elif "Approval After CM" in str(approval_seq):
            violation_reasons.append("Approval After CM Creation")
        
        # Set Violation Reason and Count
        if violation_reasons:
            df.at[i, "Violation Reason"] = " | ".join(violation_reasons)
            df.at[i, "Violation Count"] = len(violation_reasons)
        else:
            df.at[i, "Violation Reason"] = "None"
            df.at[i, "Violation Count"] = 0
        
        # SoD - Separation of Duties check
        cb = str(row["Created By"]).strip().casefold()
        apv = str(row["Approver"]).strip().casefold()
        df.at[i, "Designation Level Check"] = "Violation" if cb and apv and (cb == apv) else "OK"

# ===================== Streamlit App =====================

def main():
    # Initialize session state for chat
    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = []
    if "result_df" not in st.session_state:
        st.session_state.result_df = None
    if "validator" not in st.session_state:
        st.session_state.validator = None
    if "ai_summary_generated" not in st.session_state:
        st.session_state.ai_summary_generated = False
    
    st.markdown('<p class="header">📊 Credit Memo SOX Automation</p>', unsafe_allow_html=True)
    
    # Force cache clear on page load
    st.set_page_config(page_title="Credit Memo SOX Automation", layout="wide")
    
    # Sidebar
    with st.sidebar:
        st.header("Configuration")
        sla_days = st.number_input("SLA Threshold (Business Days)", value=5, min_value=1)
        missing_levels = st.number_input("Missing Levels for High Risk", value=2, min_value=1)
        
        # Keyword filters
        keywords_promo_input = st.text_input("Keywords – Promotional (comma-separated)", value="promotional, promotion")
        keywords_contract_input = st.text_input("Keywords – Contract (comma-separated)", value="contract")
        
        # Parse keywords
        keywords_promotional = [k.strip().lower() for k in keywords_promo_input.split(",") if k.strip()]
        keywords_contract = [k.strip().lower() for k in keywords_contract_input.split(",") if k.strip()]
        
        # New Analysis / Reset Button
        st.divider()
        if st.button("🔄 New Analysis", use_container_width=True, key="sidebar_new_analysis"):
            st.session_state.result_df = None
            st.session_state.chat_messages = []
            st.session_state.ai_summary_generated = False
            st.rerun()
    
    # Main area
    col1, col2 = st.columns([3, 1])
    
    with col1:
        uploaded_file = st.file_uploader("Upload Credit Memo Excel", type=["xlsx", "xls"])
    
    with col2:
        st.metric("Status", "Ready" if uploaded_file else "Waiting")
    
    if uploaded_file:
        st.info("📁 File uploaded. Processing...")
        
        # Check if AI credentials available (from env or Streamlit secrets)
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        
        # Try to get from secrets if not in env
        try:
            if not api_key:
                api_key = st.secrets.get("AZURE_OPENAI_API_KEY")
            if not endpoint:
                endpoint = st.secrets.get("AZURE_OPENAI_ENDPOINT")
        except:
            pass  # Secrets file doesn't exist, continue with env vars
        
        ai_available = bool(api_key and endpoint)
        
        if ai_available:
            use_ai = st.checkbox("Enable AI Analysis", value=True)
            st.success("✅ AI credentials detected - AI features enabled")
        else:
            st.warning("⚠️ AI features disabled - Add Azure credentials to .env")
            use_ai = False
        
        try:
            # Read Excel
            xl = pd.ExcelFile(uploaded_file)
            
            # Auto-detect summary sheet
            summary_sheet = next((s for s in xl.sheet_names if "summary" in s.lower()), xl.sheet_names[0])
            
            # Parse matrices
            matrices = {}
            for sheet in xl.sheet_names:
                cls = classify_matrix_sheet_name(sheet)
                if cls:
                    m = read_matrix_table(xl, sheet)
                    if m is not None:
                        matrices[cls] = m
            
            # Parse summary
            raw = xl.parse(summary_sheet, header=None)
            hdr = find_header_row(raw, ["memo", "customer", "amount"])
            if hdr is None:
                hdr = 0
            
            df = xl.parse(summary_sheet, header=hdr)
            df = norm_cols(df)
            df = map_summary_columns(df)
            
            # Data type conversions
            df["Cm Date"] = pd.to_datetime(df["Cm Date"], errors="coerce")
            df["Date Of Approval"] = pd.to_datetime(df["Date Of Approval"], errors="coerce")
            df["Amount"] = pd.to_numeric(df["Amount"].astype(str).str.replace(",", ""), errors="coerce")
            
            # Validate
            validator = ValidateCredit(sla_days, missing_levels, keywords_promotional, keywords_contract)
            result_df = validator.validate(df, matrices)
            
            # Store in session state for chat
            st.session_state.result_df = result_df
            st.session_state.validator = validator
            
            # Display summary
            st.divider()
            st.subheader("✅ Analysis Complete")
            st.markdown("Here are your validation results:")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                compliant = (result_df["SOX Status"] == "SOX Compliant").sum()
                st.metric("Compliant", compliant, f"{int(compliant/len(result_df)*100)}%")
            
            with col2:
                violations = (result_df["SOX Status"] == "SOX Violation").sum()
                st.metric("Violations", violations, f"{int(violations/len(result_df)*100)}%")
            
            with col3:
                high_risk = (result_df["Risk Level"] == "High").sum()
                med_risk = (result_df["Risk Level"] == "Medium").sum()
                st.metric("High Risk", high_risk, f"🔴 {med_risk} medium")
            
            with col4:
                total_amount = result_df["Amount"].sum()
                st.metric("Total Amount", f"${total_amount:,.0f}")
            
            # AI Insight Card
            if violations > 0 or high_risk > 0:
                insight_text = "⚠️ Found potential compliance issues requiring review."
                if violations > 0:
                    insight_text += f" {violations} violation(s) detected."
                if high_risk > 0:
                    insight_text += f" {high_risk} high-risk item(s) need attention."
                st.markdown(f"""
                <div style="background-color: #FFE5E5; border: 2px solid #FF6B6B; border-radius: 8px; padding: 16px; margin: 10px 0;">
                    <strong>🤖 AI Insight:</strong> {insight_text}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div style="background-color: #E5F5E5; border: 2px solid #51CF66; border-radius: 8px; padding: 16px; margin: 10px 0;">
                    <strong>✅ All Clear:</strong> All credit memos are compliant with SOX requirements!
                </div>
                """, unsafe_allow_html=True)
            
            # Charts
            col1, col2 = st.columns(2)
            
            with col1:
                sox_counts = result_df["SOX Status"].value_counts()
                fig, ax = plt.subplots(figsize=(6, 4))
                colors = ["#2ca02c" if x == "SOX Compliant" else "#d62728" for x in sox_counts.index]
                ax.pie(sox_counts.values, labels=sox_counts.index, autopct="%1.1f%%", colors=colors)
                ax.set_title("SOX Status Distribution")
                st.pyplot(fig)
            
            with col2:
                risk_counts = result_df["Risk Level"].value_counts()
                fig, ax = plt.subplots(figsize=(6, 4))
                colors = {"": "#808080", "Low": "#2ca02c", "Medium": "#ff7f0e", "High": "#d62728"}
                color_list = [colors.get(x, "#808080") for x in risk_counts.index]
                ax.pie(risk_counts.values, labels=risk_counts.index, autopct="%1.1f%%", colors=color_list)
                ax.set_title("Risk Level Distribution")
                st.pyplot(fig)
            
            # AI Analysis
            if use_ai and validator.ai_client:
                st.divider()
                with st.spinner("Running AI Analysis..."):
                    try:
                        prompt = f"""Analyze these credit memo validation results:
                        
Total: {len(result_df)}
Compliant: {compliant}
Violations: {violations}
High Risk: {high_risk}

Provide key findings and recommendations in 3-4 sentences."""
                        
                        response = validator.ai_client.invoke([HumanMessage(content=prompt)])
                        
                        st.subheader("🤖 AI Analysis Summary")
                        st.info(response.content)
                    except Exception as e:
                        st.warning(f"AI analysis failed: {e}")
            
            # Data table
            st.divider()
            st.subheader("Detailed Results")
            
            # Filters
            col1, col2, col3 = st.columns(3)
            
            with col1:
                memo_filter = st.text_input("Filter by Memo ID", "")
            
            with col2:
                status_filter = st.multiselect("Filter by Status", ["SOX Compliant", "SOX Violation"], default=["SOX Compliant", "SOX Violation"])
            
            with col3:
                risk_filter = st.multiselect("Filter by Risk", ["Low", "Medium", "High"], default=["Low", "Medium", "High"])
            
            # Apply filters
            filtered_df = result_df.copy()
            if memo_filter:
                filtered_df = filtered_df[filtered_df["Memo"].astype(str).str.contains(memo_filter, case=False)]
            if status_filter:
                filtered_df = filtered_df[filtered_df["SOX Status"].isin(status_filter)]
            if risk_filter:
                filtered_df = filtered_df[filtered_df["Risk Level"].isin(risk_filter)]
            
            # Display columns - same as credit.py
            display_cols = [
                "Memo", "Customer Name", "Date Of Approval", "Cm Date", "Reason", "Reason Class", "Amount",
                "Approver", "Approver Designation",
                "Required Approval Level", "Final Approver Level", "Final Approver",
                "SOX Status", "Risk Level", "Violation Reason", "Missing Approvals",
                "Approval Timeline (Business Days)", "Timeline Status", "Approval Sequence",
                "Designation Level Check", "Duplicate Memo"
            ]
            
            have_cols = [c for c in display_cols if c in filtered_df.columns]
            
            # Format the dataframe for display
            display_df = filtered_df[have_cols].copy()
            # Reset index to start fresh (remove pandas index)
            display_df = display_df.reset_index(drop=True)
            # Add row numbers starting from 1
            display_df.insert(0, "Row #", range(1, len(display_df) + 1))
            display_df["Cm Date"] = display_df["Cm Date"].astype(str).str.replace(" 00:00:00", "")
            display_df["Date Of Approval"] = display_df["Date Of Approval"].astype(str).str.replace(" 00:00:00", "")
            display_df["Amount"] = display_df["Amount"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
            
            # Define styling function
            def style_cell(val):
                """Style cells based on values"""
                if val == "SOX Compliant":
                    return 'background-color: #C6EFCE; color: black; font-weight: bold'
                elif val == "SOX Violation":
                    return 'background-color: #FFC7CE; color: black; font-weight: bold'
                elif val == "High":
                    return 'background-color: #FFC7CE; color: black; font-weight: bold'
                elif val == "Medium":
                    return 'background-color: #F8CBAD; color: black'
                else:
                    return ''
            
            def style_violation_reason(val):
                """Style Violation Reason column based on violation count"""
                if pd.isna(val) or val == "None" or str(val).strip().lower() == "none":
                    return 'background-color: #E7E6E6; color: black'  # Gray for no violation
                
                # Count the number of violations (separated by |)
                reason_str = str(val)
                violation_count = reason_str.count('|') + 1 if reason_str.strip() else 0
                
                if violation_count == 1:
                    return 'background-color: #FF6B6B; color: white; font-weight: bold'  # Red for 1 violation
                elif violation_count >= 2:
                    return 'background-color: #4A90E2; color: white; font-weight: bold'  # Blue for 2+ violations
                else:
                    return 'background-color: #E7E6E6; color: black'
            
            # Display styled table
            styled_df = display_df.style.map(style_cell, subset=["SOX Status", "Risk Level"])
            
            # Only apply Violation Reason styling if the column exists
            if "Violation Reason" in display_df.columns:
                styled_df = styled_df.map(style_violation_reason, subset=["Violation Reason"])
            
            st.dataframe(
                styled_df,
                width='stretch',
                height=600,
                hide_index=True
            )
            
            # ===== Quick Action AI Buttons =====
            st.divider()
            st.subheader("🤖 Quick AI Insights")
            
            col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
            
            with col_btn1:
                if st.button("📋 Explain Violations", use_container_width=True):
                    violations_df = result_df[result_df["SOX Status"] == "SOX Violation"]
                    if not violations_df.empty:
                        query = "Summarize the key violations found and explain their impact on SOX compliance."
                        context = build_context_prompt(result_df)
                        response = get_ai_response(st.session_state.validator.ai_client, context, query)
                        st.session_state.chat_messages.append({"role": "user", "content": query})
                        st.session_state.chat_messages.append({"role": "assistant", "content": response})
                        st.info(f"✅ Added to AI chat: {query}")
                    else:
                        st.success("✅ No violations found!")
            
            with col_btn2:
                if st.button("⚠️ High-Risk Items", use_container_width=True):
                    high_risk_df = result_df[result_df["Risk Level"] == "High"]
                    if not high_risk_df.empty:
                        query = "What are the high-risk items and what immediate actions should be taken?"
                        context = build_context_prompt(result_df)
                        response = get_ai_response(st.session_state.validator.ai_client, context, query)
                        st.session_state.chat_messages.append({"role": "user", "content": query})
                        st.session_state.chat_messages.append({"role": "assistant", "content": response})
                        st.info(f"✅ Added to AI chat: {query}")
                    else:
                        st.success("✅ No high-risk items!")
            
            with col_btn3:
                if st.button("⏱️ Timeline Analysis", use_container_width=True):
                    sla_violations = result_df[result_df["Timeline Status"].astype(str).str.startswith("Over")]
                    query = "Analyze the timeline violations and suggest improvements to meet SLA requirements."
                    context = build_context_prompt(result_df)
                    response = get_ai_response(st.session_state.validator.ai_client, context, query)
                    st.session_state.chat_messages.append({"role": "user", "content": query})
                    st.session_state.chat_messages.append({"role": "assistant", "content": response})
                    st.info(f"✅ Added to AI chat: {query}")
            
            with col_btn4:
                if st.button("📊 Generate Summary", use_container_width=True):
                    query = "Provide an executive summary of the validation results including key metrics, critical issues, and top 3 recommendations."
                    context = build_context_prompt(result_df)
                    response = get_ai_response(st.session_state.validator.ai_client, context, query)
                    st.session_state.chat_messages.append({"role": "user", "content": query})
                    st.session_state.chat_messages.append({"role": "assistant", "content": response})
                    st.info(f"✅ Added to AI chat: {query}")
            
            # Quick download option in preview
            st.write("")  # Small space
            download_col = st.columns([1, 4])[0]  # Left align
            with download_col:
                try:
                    # Prepare full results CSV
                    csv_quick_df = result_df[have_cols].copy()
                    csv_quick_df = csv_quick_df.reset_index(drop=True)
                    csv_quick_df.insert(0, "Row #", range(1, len(csv_quick_df) + 1))
                    csv_quick_df["Cm Date"] = csv_quick_df["Cm Date"].astype(str).str.replace(" 00:00:00", "")
                    csv_quick_df["Date Of Approval"] = csv_quick_df["Date Of Approval"].astype(str).str.replace(" 00:00:00", "")
                    csv_quick_df["Amount"] = csv_quick_df["Amount"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
                    
                    csv_data = csv_quick_df.to_csv(index=False)
                    csv_file = f"credit_memo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    st.download_button(
                        label="⬇️ Quick CSV Download",
                        data=csv_data,
                        file_name=csv_file,
                        mime="text/csv",
                        use_container_width=False
                    )
                except Exception as e:
                    st.error(f"CSV download error: {e}")
            
            # Export
            st.divider()
            st.subheader("📥 Export Results")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # CSV export - FULL RESULTS (not filtered preview)
                try:
                    # Prepare export dataframe with all data
                    csv_export_df = result_df[have_cols].copy()
                    csv_export_df = csv_export_df.reset_index(drop=True)
                    csv_export_df.insert(0, "Row #", range(1, len(csv_export_df) + 1))
                    csv_export_df["Cm Date"] = csv_export_df["Cm Date"].astype(str).str.replace(" 00:00:00", "")
                    csv_export_df["Date Of Approval"] = csv_export_df["Date Of Approval"].astype(str).str.replace(" 00:00:00", "")
                    csv_export_df["Amount"] = csv_export_df["Amount"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
                    
                    csv_data = csv_export_df.to_csv(index=False)
                    csv_file = f"credit_memo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    st.download_button(
                        label="📋 Download CSV",
                        data=csv_data,
                        file_name=csv_file,
                        mime="text/csv"
                    )
                except Exception as e:
                    st.error(f"CSV export error: {e}")
            
            with col2:
                # Excel export - FULL RESULTS (not filtered preview)
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Credit Memo Check Results"
                    
                    # Export the same columns as display, with row numbers - using FULL DATA
                    excel_export_df = result_df[have_cols].copy()
                    excel_export_df = excel_export_df.reset_index(drop=True)
                    excel_export_df.insert(0, "Row #", range(1, len(excel_export_df) + 1))
                    excel_export_df["Cm Date"] = excel_export_df["Cm Date"].astype(str).str.replace(" 00:00:00", "")
                    excel_export_df["Date Of Approval"] = excel_export_df["Date Of Approval"].astype(str).str.replace(" 00:00:00", "")
                    excel_export_df["Amount"] = excel_export_df["Amount"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
                    excel_export_df = excel_export_df.where(pd.notna(excel_export_df), None)
                    
                    # Write data
                    for r in dataframe_to_rows(excel_export_df, index=False, header=True):
                        ws.append(r)
                    
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions
                    
                    # Color coding - match preview using original result_df for accurate violation counts
                    cols = list(excel_export_df.columns)
                    sox_idx = cols.index("SOX Status") + 1 if "SOX Status" in cols else None
                    violation_reason_idx = cols.index("Violation Reason") + 1 if "Violation Reason" in cols else None
                    violation_count_idx = cols.index("Violation Count") + 1 if "Violation Count" in cols else None
                    risk_idx = cols.index("Risk Level") + 1 if "Risk Level" in cols else None
                    
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thin = Side(style="thin", color="DDDDDD")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    
                    # Color definitions
                    GREEN = "C6EFCE"
                    RED_HIGHLIGHT = "FF6B6B"
                    BLUE_HIGHLIGHT = "4A90E2"
                    GREY = "E7E6E6"
                    
                    # Format header
                    for cell in ws[1]:
                        cell.font = Font(bold=True, size=11, color="000000")
                        cell.alignment = center
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # Format data rows - Match preview styling
                    for r in range(2, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.alignment = center
                            cell.border = border
                            cell.font = Font(size=10, color="000000")
                        
                        # Color Violation Reason cell - MATCH PREVIEW EXACTLY
                        if violation_reason_idx:
                            reason_val = ws.cell(row=r, column=violation_reason_idx).value
                            reason_str = str(reason_val) if reason_val else ""
                            
                            # Get violation count from Violation Count column if available
                            violation_count = 0
                            if violation_count_idx:
                                count_val = ws.cell(row=r, column=violation_count_idx).value
                                try:
                                    violation_count = int(count_val) if count_val else 0
                                except:
                                    violation_count = 0
                            else:
                                # Fallback: count pipes in reason string
                                if reason_str.strip().lower() != "none" and reason_str.strip():
                                    violation_count = reason_str.count('|') + 1
                            
                            # Apply color based on violation count
                            if violation_count == 0 or reason_str.strip().lower() == "none":
                                # Gray for no violation (matching preview)
                                ws.cell(row=r, column=violation_reason_idx).fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
                                ws.cell(row=r, column=violation_reason_idx).font = Font(color="000000", size=10)
                            elif violation_count == 1:
                                # RED for exactly 1 violation (matching preview #FF6B6B)
                                ws.cell(row=r, column=violation_reason_idx).fill = PatternFill(start_color=RED_HIGHLIGHT, end_color=RED_HIGHLIGHT, fill_type="solid")
                                ws.cell(row=r, column=violation_reason_idx).font = Font(color="FFFFFF", bold=True, size=10)
                            elif violation_count >= 2:
                                # BLUE for 2+ violations (matching preview #4A90E2)
                                ws.cell(row=r, column=violation_reason_idx).fill = PatternFill(start_color=BLUE_HIGHLIGHT, end_color=BLUE_HIGHLIGHT, fill_type="solid")
                                ws.cell(row=r, column=violation_reason_idx).font = Font(color="FFFFFF", bold=True, size=10)
                    
                    # Auto column widths
                    for col_idx in range(1, ws.max_column + 1):
                        max_len = max(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1))
                        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 50)
                    
                    excel_file = f"credit_memo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    wb.save(excel_file)
                    
                    with open(excel_file, "rb") as f:
                        st.download_button(
                            label="📊 Download Excel",
                            data=f.read(),
                            file_name=excel_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Excel export error: {e}")
            
            with col3:
                # PDF export
                try:
                    from fpdf import FPDF
                    import tempfile
                    
                    # Create charts
                    def create_chart(fig_title, create_func):
                        fig, ax = create_func()
                        fig.suptitle(fig_title, fontsize=12, fontweight='bold')
                        # Save to temp file
                        temp_img = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                        fig.savefig(temp_img.name, dpi=150, bbox_inches='tight')
                        plt.close(fig)
                        return temp_img.name
                    
                    chart_images = []
                    
                    # SOX Status pie chart
                    def create_sox_chart():
                        fig, ax = plt.subplots(figsize=(8, 6))
                        result_df["SOX Status"].value_counts().plot(kind='pie', autopct='%1.1f%%', ax=ax)
                        ax.set_ylabel('')
                        return fig, ax
                    chart_images.append(("SOX Status Distribution", create_chart("SOX Status", create_sox_chart)))
                    
                    # Risk Level pie chart
                    def create_risk_chart():
                        fig, ax = plt.subplots(figsize=(8, 6))
                        result_df["Risk Level"].replace("", pd.NA).value_counts().plot(kind='pie', autopct='%1.1f%%', ax=ax)
                        ax.set_ylabel('')
                        return fig, ax
                    chart_images.append(("Risk Level Distribution", create_chart("Risk Level", create_risk_chart)))
                    
                    # Timeline Status pie chart
                    def create_timeline_chart():
                        fig, ax = plt.subplots(figsize=(8, 6))
                        result_df["Timeline Status"].value_counts().plot(kind='pie', autopct='%1.1f%%', ax=ax)
                        ax.set_ylabel('')
                        return fig, ax
                    chart_images.append(("Timeline Status", create_chart("Timeline Status", create_timeline_chart)))
                    
                    # Violations by Approver
                    def create_approver_chart():
                        fig, ax = plt.subplots(figsize=(10, 6))
                        violations = result_df[result_df["SOX Status"] == "SOX Violation"]["Approver Designation"].value_counts().head(10)
                        violations.plot(kind='barh', ax=ax)
                        ax.set_xlabel('Count')
                        ax.set_title('Top Approvers with Violations')
                        return fig, ax
                    if (result_df["SOX Status"] == "SOX Violation").any():
                        chart_images.append(("Violations by Approver", create_chart("Top Violations", create_approver_chart)))
                    
                    # Create PDF
                    pdf = FPDF()
                    pdf.set_auto_page_break(auto=True, margin=15)
                    
                    # Cover page
                    pdf.add_page()
                    pdf.set_font("helvetica", "B", 16)
                    pdf.cell(0, 15, "Credit Memo SOX Audit Report", new_x="LMARGIN", new_y="NEXT", align="C")
                    pdf.ln(5)
                    pdf.set_font("helvetica", size=11)
                    
                    compliant = (result_df["SOX Status"] == "SOX Compliant").sum()
                    violations = (result_df["SOX Status"] == "SOX Violation").sum()
                    high_risk = (result_df["Risk Level"] == "High").sum()
                    med_risk = (result_df["Risk Level"] == "Medium").sum()
                    
                    pdf.multi_cell(0, 7,
                        f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                        f"Total Memos Reviewed: {len(result_df)}\n"
                        f"Compliant: {compliant} ({100*compliant/len(result_df):.1f}%)\n"
                        f"Violations: {violations} ({100*violations/len(result_df):.1f}%)\n\n"
                        f"Risk Levels:\n"
                        f"  High Risk: {high_risk}\n"
                        f"  Medium Risk: {med_risk}")
                    
                    # Add chart pages
                    for chart_title, chart_path in chart_images:
                        pdf.add_page()
                        pdf.set_font("helvetica", "B", 12)
                        pdf.cell(0, 8, chart_title, new_x="LMARGIN", new_y="NEXT")
                        pdf.ln(3)
                        if os.path.exists(chart_path):
                            pdf.image(chart_path, x=15, y=pdf.get_y(), w=180)
                            os.unlink(chart_path)  # Clean up temp file
                    
                    pdf_file = f"credit_memo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                    pdf.output(pdf_file)
                    
                    with open(pdf_file, "rb") as f:
                        st.download_button(
                            label="📄 Download PDF (with Charts)",
                            data=f.read(),
                            file_name=pdf_file,
                            mime="application/pdf"
                        )
                except Exception as e:
                    st.error(f"PDF export error: {e}")
                    st.rerun()
        
        except Exception as e:
            st.error(f"Error processing file: {e}")
            st.write(str(e))
    
    # ===== MAIN SCREEN AI CHAT INTERFACE =====
    if "show_chat_history" not in st.session_state:
        st.session_state.show_chat_history = False
    
    if st.session_state.result_df is not None:
        if st.session_state.validator and st.session_state.validator.ai_client:
            st.divider()
            st.subheader("� AI Assistant - Ask About Results")
            
            # Chat input
            if user_query := st.chat_input("Ask about validation results..."):
                # Add user message
                st.session_state.chat_messages.append({"role": "user", "content": user_query})
                
                with st.chat_message("user", avatar="👤"):
                    st.markdown(user_query)
                
                # Generate response
                if st.session_state.result_df is not None:
                    with st.chat_message("assistant", avatar="🤖"):
                        response_placeholder = st.empty()
                        response_placeholder.markdown("🤔 Thinking...")
                        
                        context = build_context_prompt(st.session_state.result_df)
                        response = get_ai_response(st.session_state.validator.ai_client, context, user_query)
                        
                        response_placeholder.markdown(response)
                        st.session_state.chat_messages.append({"role": "assistant", "content": response})
                else:
                    st.warning("Please validate data first")
            
            # Show chat history if toggled on
            if st.session_state.show_chat_history and st.session_state.chat_messages:
                st.markdown("---")
                st.subheader("📜 Chat History")
                for msg in st.session_state.chat_messages:
                    role = msg["role"]
                    avatar = "👤" if role == "user" else "🤖"
                    with st.chat_message(role, avatar=avatar):
                        st.markdown(msg["content"], unsafe_allow_html=True)
            
            # Control buttons
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🗑️ Clear Chat"):
                    st.session_state.chat_messages = []
                    st.session_state.show_chat_history = False
                    st.rerun()
            with col2:
                history_label = "📜 Hide History" if st.session_state.show_chat_history else "📜 Show History"
                if st.button(history_label):
                    st.session_state.show_chat_history = not st.session_state.show_chat_history
                    st.rerun()
        else:
            st.warning("⚠️ AI Assistant not available. Check credentials.")
    
    # ===== SIDEBAR CHAT HISTORY LINK =====
    if st.session_state.result_df is not None:
        with st.sidebar:
            st.divider()
            if st.session_state.validator and st.session_state.validator.ai_client:
                st.subheader("💬 Chat")
                if st.button("📖 View Chat History", use_container_width=True):
                    st.session_state.show_chat_history = not st.session_state.show_chat_history
                    st.rerun()
                
                if st.session_state.chat_messages:
                    st.caption(f"💬 {len(st.session_state.chat_messages)} messages")
                else:
                    st.caption("No chat messages yet")
            else:
                st.warning("⚠️ AI not available")


if __name__ == "__main__":
    main()
